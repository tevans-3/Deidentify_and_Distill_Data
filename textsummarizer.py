# -*- coding: utf-8 -*-
"""
Anonymizing sensitive data and using Gemini to summarize and analyze student evaluations
"""

import openpyxl
from openpyxl import Workbook
import spacy
import re
import google.generativeai as genai
genai.configure(api_key="AIzaSyBTAu05pQbGou5vt2Vww5Otkhy-AjjW5js")

import ExtractedText
from ExtractedText import ExtractedText
import SpacyNLP

spreadsheet_filename = 'UNREDACTED.xlsx'

def generate_redacted(spreadsheet_filename):
  unredacted_text = ExtractedText(spreadsheet_filename)
  redacted_text = unredacted_text.redact()
  new_spreadsheet = Workbook()
  new_worksheet = new_spreadsheet.active 
  new_worksheet.title = 'Redacted'
  #add headers
  new_worksheet.append(unredacted_text.worksheet_headers)
  #next_row = 2
  
  num_columns = len(unredacted_text.worksheet_headers)
  index = 0
  #construct new spreadsheet with the formatting of the old one
  for row in unredacted_text.worksheet.iter_rows(min_row=2): 
    for cell in row: 
      if cell.value != None:
        new_worksheet.cell(row=cell.row, column=cell.col_idx).value = redacted_text[index]
        index += 1

  new_spreadsheet.save('REDACTED.xlsx')
  return redacted_text

redacted_text = generate_redacted(spreadsheet_filename)
#model = genai.GenerativeModel("gemini-1.5-flash")
#response = model.generate_content(f'The following text contains evaluations of medical preceptors. It was completed anonymously by students. First, please summarize the feedback received for each preceptor. Draw out themes and highlights. At the end, provide a precise sentiment analysis of the feedback. {redacted_text}.')
#print(response.text)